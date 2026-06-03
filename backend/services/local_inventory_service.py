"""
本地仓库存服务 - 导入/查询/统计
"""
import io
import logging
import os
import tempfile
from datetime import date, datetime
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from models.local_inventory import LocalInventory
from models.restock import InventorySnapshot

logger = logging.getLogger(__name__)

# Excel 列名映射（支持多种常见列名）
LOCAL_INV_FIELD_MAPPING = {
    "ASIN": "asin",
    "asin": "asin",
    "Asin": "asin",
    "SKU": "sku",
    "sku": "sku",
    "Sku": "sku",
    "品名": "product_name",
    "商品名称": "product_name",
    "产品名称": "product_name",
    "Product Name": "product_name",
    "product_name": "product_name",
    "店铺": "account",
    "账号": "account",
    "Account": "account",
    "account": "account",
    "国家": "country",
    "国家（地区）": "country",
    "站点": "country",
    "Country": "country",
    "country": "country",
    "本地库存": "quantity",
    "本地仓库存": "quantity",
    "本地可用": "quantity",
    "库存数量": "quantity",
    "数量": "quantity",
    "Quantity": "quantity",
    "quantity": "quantity",
}


def import_local_inventory(db: Session, file_content: bytes, filename: str = None) -> dict:
    """
    导入本地仓库存Excel数据
    Excel格式要求：至少包含 ASIN/SKU 和 库存数量 列
    """
    df = pd.read_excel(io.BytesIO(file_content))
    total_rows = len(df)
    logger.info(f"本地仓库存Excel读取完成: {total_rows} 条")

    # 重命名列
    df = df.rename(columns={k: v for k, v in LOCAL_INV_FIELD_MAPPING.items() if k in df.columns})

    # 检查必要列
    if "asin" not in df.columns and "sku" not in df.columns:
        raise ValueError("Excel中未找到 ASIN 或 SKU 列，请检查表头")
    if "quantity" not in df.columns:
        raise ValueError("Excel中未找到库存数量列（支持列名：本地库存/本地仓库存/库存数量/数量/Quantity）")

    # 数据清洗
    today = date.today()

    for col in ["asin", "sku", "product_name", "account", "country"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)

    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0)
    df["tenant_id"] = 1
    df["batch_date"] = today

    # 清理旧数据
    db.execute(text("DELETE FROM local_inventories WHERE tenant_id = 1"))
    db.commit()
    logger.info("本地仓旧数据清理完成")

    # 插入新数据
    insert_cols = [c for c in ["asin", "sku", "product_name", "account", "country", "quantity", "tenant_id", "batch_date"] if c in df.columns]
    df_to_insert = df[insert_cols].copy()

    df_to_insert.to_sql(
        "local_inventories",
        db.get_bind(),
        if_exists="append",
        index=False,
        chunksize=5000,
        method="multi",
    )
    logger.info(f"本地仓库存导入完成: {total_rows} 条")

    # 统计
    total_qty = float(df["quantity"].sum())
    unique_asin = df["asin"].nunique() if "asin" in df.columns else 0

    return {
        "total_rows": total_rows,
        "total_quantity": total_qty,
        "unique_asin": unique_asin,
        "batch_date": today.isoformat(),
    }


def get_local_inventory_summary(db: Session) -> dict:
    """获取本地仓库存汇总"""
    total_sku = db.query(func.count(LocalInventory.id)).filter(LocalInventory.tenant_id == 1).scalar() or 0
    total_qty = db.query(func.sum(LocalInventory.quantity)).filter(LocalInventory.tenant_id == 1).scalar() or 0
    latest_batch = db.query(func.max(LocalInventory.batch_date)).filter(LocalInventory.tenant_id == 1).scalar()

    return {
        "total_sku": total_sku,
        "total_quantity": float(total_qty),
        "latest_batch_date": latest_batch.isoformat() if latest_batch else None,
    }


def get_local_inventory_list(db: Session, keyword: str = None, page: int = 1, page_size: int = 20) -> dict:
    """查询本地仓库存列表"""
    query = db.query(LocalInventory).filter(LocalInventory.tenant_id == 1)

    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(
            (LocalInventory.asin.like(kw)) |
            (LocalInventory.sku.like(kw)) |
            (LocalInventory.product_name.like(kw)) |
            (LocalInventory.account.like(kw))
        )

    total = query.count()
    items = query.order_by(LocalInventory.asin).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "items": [{
            "id": item.id,
            "asin": item.asin,
            "sku": item.sku,
            "product_name": item.product_name,
            "account": item.account,
            "country": item.country,
            "quantity": float(item.quantity) if item.quantity else 0,
            "batch_date": item.batch_date.isoformat() if item.batch_date else None,
        } for item in items],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


def get_local_inventory_by_asin(db: Session, asin: str, account: str = None) -> float:
    """根据ASIN查询本地仓库存数量"""
    query = db.query(func.sum(LocalInventory.quantity)).filter(
        LocalInventory.tenant_id == 1,
        LocalInventory.asin == asin,
    )
    if account:
        query = query.filter(LocalInventory.account == account)
    result = query.scalar()
    return float(result) if result else 0


def get_local_inventory_map(db: Session) -> dict:
    """获取所有本地仓库存映射 {(asin, account): quantity}"""
    items = db.query(LocalInventory.asin, LocalInventory.account, LocalInventory.quantity).filter(
        LocalInventory.tenant_id == 1
    ).all()

    inv_map = {}
    for item in items:
        key = (item.asin or "", item.account or "")
        inv_map[key] = float(item.quantity) if item.quantity else 0
    return inv_map


def clear_local_inventory(db: Session) -> dict:
    """清空本地仓库存数据"""
    result = db.execute(text("DELETE FROM local_inventories WHERE tenant_id = 1"))
    db.commit()
    return {"deleted_count": result.rowcount}


def import_reduction_table(db: Session, country: str, file_content: bytes) -> dict:
    """
    导入已采购数据（减量表）
    根据Excel中的SKU匹配库存快照FNSKU，将已采购数量写入local_inventories表
    """
    logger.info(f"减量表导入开始, country={country}")

    sku_columns = {"SKU", "sku", "Sku"}
    purchase_columns = {"已采购", "已采购数", "采购数量"}

    # 尝试读取 Excel，有些文件第一行是合并标题（如"C美国"），实际列头在第二行
    df = pd.read_excel(io.BytesIO(file_content))
    total_rows = len(df)
    logger.info(f"减量表Excel读取完成: {total_rows} 条")

    sku_col = None
    purchase_col = None
    for col in df.columns:
        if col.strip() in sku_columns:
            sku_col = col
        if col.strip() in purchase_columns:
            purchase_col = col

    # 如果未找到关键列，尝试跳过第一行重新读取（处理合并标题行）
    if not sku_col or not purchase_col:
        logger.warning("未在首行找到标准列名，尝试跳过第一行重新读取")
        df2 = pd.read_excel(io.BytesIO(file_content), header=1)
        for col in df2.columns:
            if col.strip() in sku_columns:
                sku_col = col
            if col.strip() in purchase_columns:
                purchase_col = col
        if sku_col and purchase_col:
            df = df2
            total_rows = len(df)
            logger.info(f"跳过首行后重新读取: {total_rows} 条")

    # 最终回退：仍找不到SKU列则使用第一列
    if not sku_col:
        sku_col = df.columns[0]
        logger.warning(f"未找到标准SKU列名，默认使用第一列 '{sku_col}' 作为SKU列")

    if not purchase_col:
        raise ValueError("Excel中未找到已采购列（支持列名：已采购/已采购数/采购数量）")

    latest_date = db.query(func.max(InventorySnapshot.snapshot_date)).scalar()
    today = date.today()

    updated = 0
    skipped = 0
    updated_ids = []

    results = []

    for idx, row in df.iterrows():
        sku_val = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""
        purchase_val = row[purchase_col]

        if not sku_val:
            results.append({"sku": "", "status": "失败", "reason": "SKU为空", "asin": "", "product_name": "", "account": "", "quantity": ""})
            skipped += 1
            continue

        try:
            qty = float(purchase_val) if pd.notna(purchase_val) else None
            if qty is None:
                results.append({"sku": sku_val, "status": "失败", "reason": "已采购数量为空", "asin": "", "product_name": "", "account": "", "quantity": ""})
                skipped += 1
                continue
        except (ValueError, TypeError):
            results.append({"sku": sku_val, "status": "失败", "reason": "已采购数量非数值", "asin": "", "product_name": "", "account": "", "quantity": str(purchase_val)})
            skipped += 1
            continue

        snapshot = None
        if latest_date:
            snapshot = db.query(InventorySnapshot).filter(
                InventorySnapshot.snapshot_date == latest_date,
                InventorySnapshot.msku == sku_val,
                InventorySnapshot.country == country,
                InventorySnapshot.deleted_at.is_(None),
            ).first()

        if not snapshot:
            results.append({"sku": sku_val, "status": "失败", "reason": "未找到匹配的 MSKU", "asin": "", "product_name": "", "account": "", "quantity": str(qty)})
            skipped += 1
            continue

        previous_val = snapshot.local_inventory or 0
        snapshot.local_inventory = qty
        snapshot.total_stock = (snapshot.fba_stock or 0) + (snapshot.fba_inbound or 0) + qty - (snapshot.inspection_quantity or 0)
        updated += 1
        updated_ids.append(snapshot.id)
        results.append({"sku": sku_val, "status": "成功", "reason": f"已更新（原值: {previous_val} → {qty}）", "asin": snapshot.asin or "", "product_name": snapshot.product_name or "", "account": snapshot.account or "", "quantity": str(qty)})

    db.commit()
    logger.info(f"减量表导入完成: 更新{updated}, 跳过{skipped}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导入结果"

    headers = ["SKU", "状态", "失败原因", "ASIN", "品名", "店铺", "已采购数量"]
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result["sku"])
        ws.cell(row=row_idx, column=2, value=result["status"])
        ws.cell(row=row_idx, column=3, value=result["reason"])
        ws.cell(row=row_idx, column=4, value=result["asin"])
        ws.cell(row=row_idx, column=5, value=result["product_name"])
        ws.cell(row=row_idx, column=6, value=result["account"])
        ws.cell(row=row_idx, column=7, value=result["quantity"])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reduction_result_{timestamp}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)
    logger.info(f"减量表导入结果已保存: {filepath}")

    return {
        "total": total_rows,
        "updated": updated,
        "skipped": skipped,
        "snapshot_ids": updated_ids,
        "result_file_id": filename,
    }
